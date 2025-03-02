import random  # Import random for generating random values
import pandas as pd  # Import pandas for data manipulation
from openpyxl import load_workbook  # Import load_workbook for working with Excel files
from openpyxl.styles import PatternFill, Font  # Import PatternFill and Font for styling Excel cells
import os  # Import os for file operations
from collections import Counter  # Import Counter for counting occurrences of elements

print("Script execution started.")  # Ensure the script is being executed

def generate_unique_numbers(count, start, end):
    """
    Generate a list of unique random numbers.

    Parameters:
    count (int): Number of random numbers to generate.
    start (int): The lower bound of the random numbers.
    end (int): The upper bound of the random numbers.

    Returns:
    list: A list of unique random numbers.
    """
    return random.sample(range(start, end + 1), count)  # Generate a list of unique random numbers

def generate_numbers(count, start, end, columns, max_repeats):
    """
    Generate a list of random numbers with constraints.

    Parameters:
    count (int): Number of random numbers to generate per column.
    start (int): The lower bound of the random numbers.
    end (int): The upper bound of the random numbers.
    columns (int): Number of columns.
    max_repeats (float): Maximum allowed frequency of repetition.

    Returns:
    list: A list of random numbers.
    """
    numbers = []  # Initialize an empty list to store the numbers
    all_numbers = []  # Initialize an empty list to store all generated numbers
    for _ in range(columns):  # Generate numbers for the specified number of columns
        col_numbers = generate_unique_numbers(count, start, end)  # Generate unique numbers for the column
        numbers.append(col_numbers)  # Append the column numbers to the list
        all_numbers.extend(col_numbers)  # Extend the list of all numbers with the column numbers
    
    total_numbers = count * columns  # Calculate the total number of generated numbers
    max_repeats = total_numbers * max_repeats  # Calculate the maximum allowed repeats
    number_counts = Counter(all_numbers)  # Count the occurrences of each number
    for number, freq in number_counts.items():  # Iterate through the counted numbers
        if freq > max_repeats:  # If the frequency of a number exceeds the maximum allowed repeats
            excess = freq - max_repeats  # Calculate the excess repeats
            for _ in range(int(excess)):  # Remove the excess repeats
                for col in numbers:
                    if number in col:
                        col.remove(number)
                        new_number = random.choice([n for n in range(start, end + 1) if n not in all_numbers])
                        col.append(new_number)
                        all_numbers.append(new_number)
                        break
    
    common_numbers = random.sample(range(start, end + 1), int(count * 0.1))  # Ensure some numbers are present in all columns
    for number in common_numbers:
        for col in numbers:
            if number not in col:
                col[random.randint(0, count - 1)] = number
    
    for number in common_numbers:
        for col in numbers:
            if number not in col:
                col[random.randint(0, count - 1)] = number
    
    return numbers  # Return the generated numbers

def save_to_excel(numbers, filename):
    """
    Save the generated numbers to an Excel file.

    Parameters:
    numbers (list): The list of generated numbers.
    filename (str): The name of the Excel file.
    """
    df = pd.DataFrame(numbers).transpose()  # Create a DataFrame from the numbers and transpose it
    df.columns = [f"Nr{i+1}" for i in range(len(numbers))]  # Set the column names
    df.to_excel(filename, index=False)  # Save the DataFrame to an Excel file
    print(f"Numbers saved to {filename}")  # Print a message indicating the file was saved
    color_cells(filename, numbers)  # Call the function to color the cells

def save_to_csv(numbers, filename):
    """
    Save the generated numbers to a CSV file.

    Parameters:
    numbers (list): The list of generated numbers.
    filename (str): The name of the CSV file.
    """
    df = pd.DataFrame(numbers).transpose()  # Create a DataFrame from the numbers and transpose it
    df.columns = [f"Nr{i+1}" for i in range(len(numbers))]  # Set the column names
    df.to_csv(filename, index=False)  # Save the DataFrame to a CSV file
    print(f"Numbers saved to {filename}")  # Print a message indicating the file was saved

def color_cells(filename, numbers):
    """
    Color the cells in the Excel file with random colors for repeated numbers.

    Parameters:
    filename (str): The name of the Excel file.
    numbers (list): The list of generated numbers.
    """
    wb = load_workbook(filename)  # Load the Excel workbook
    ws = wb.active  # Get the active worksheet
    number_counts = Counter([num for sublist in numbers for num in sublist])  # Count the occurrences of each number
    repeated_numbers = {num: count for num, count in number_counts.items() if count > 1}  # Filter out non-repeated numbers
    color_map = {num: random_color() for num in repeated_numbers}  # Assign a random color to each repeated number
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):  # Iterate through the cells
        for cell in row:
            if cell.value in repeated_numbers:  # If the cell value is a repeated number
                bg_color = color_map[cell.value]  # Get the assigned color for the number
                cell.fill = PatternFill(start_color=f"FF{bg_color[1:]}", end_color=f"FF{bg_color[1:]}", fill_type="solid")  # Set the cell fill color
                cell.font = Font(color="000000")  # Keep the text color black
    wb.save(filename)  # Save the workbook
    print(f"Cells colored in {filename}")  # Print a message indicating the cells were colored

def random_color():
    """
    Generate a random color in hexadecimal format.

    Returns:
    str: A random color in hexadecimal format.
    """
    return f"#{random.randint(0, 0xFFFFFF):06X}"  # Generate a random color

def main():
    """
    Main function to generate numbers and save them to files.
    """
    try:
        print("This is a simple text to ensure printing works.")  # Print a simple text to ensure printing works
        
        count = int(input("Enter the number of rows: "))  # Number of random numbers to generate per column
        start = int(input("Enter the lower bound of the random numbers: "))  # Lower bound of the random numbers
        end = int(input("Enter the upper bound of the random numbers: "))  # Upper bound of the random numbers
        columns = int(input("Enter the number of columns: "))  # Number of columns
        max_repeats = float(input("Enter the maximum allowed frequency of repetition (e.g., 0.05 for 5%): "))  # Maximum allowed frequency of repetition
        
        numbers = generate_numbers(count, start, end, columns, max_repeats)  # Generate the numbers
        print("Generated numbers (first column):", numbers[0][:10], "...")  # Print first 10 numbers of the first column for brevity
        save_to_excel(numbers, "numbers.xlsx")  # Save the numbers to an Excel file
        save_to_csv(numbers, "numbers.csv")  # Save the numbers to a CSV file
    except Exception as e:
        print(f"An error occurred: {e}")  # Print an error message if an exception occurs

if __name__ == "__main__":
    main()  # Call the main function if the script is executed directly
    print("Script execution finished.")  # Ensure the script finishes execution

    script_path = os.path.abspath(__file__)  # Get the absolute path of the script
    with open(script_path, 'r') as file:
        script_content = file.read()  # Read the script content
    with open(script_path, 'w') as file:
        file.write(script_content)  # Write the script content back to the file
    print(f"Script saved automatically to {script_path}")  # Print a message indicating the script was saved
